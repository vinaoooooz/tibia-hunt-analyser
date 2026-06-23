#!/usr/bin/env python3
"""
Tibia Hunt Analyzer Excel Generator
Parses N hunt analyzer sessions (TXT or JSON), calculates metrics, generates Excel
with creature images from TibiaWiki.
"""

import re
import os
import io
import sys
import json
import argparse
import urllib.parse
from datetime import datetime
from statistics import mean
from pathlib import Path

import requests
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart


CREATURE_IMAGES = {
    "arachnophobica": "https://static.wikia.nocookie.net/tibia/images/2/21/Arachnophobica.gif/revision/latest?cb=20181107170955&path-prefix=en",
    "burster spectre": "https://static.wikia.nocookie.net/tibia/images/1/1f/Burster_Spectre.gif/revision/latest?cb=20181107180045&path-prefix=en",
    "gazer spectre": "https://static.wikia.nocookie.net/tibia/images/8/82/Gazer_Spectre.gif/revision/latest?cb=20181107180045&path-prefix=en",
    "ripper spectre": "https://static.wikia.nocookie.net/tibia/images/6/6c/Ripper_Spectre.gif/revision/latest?cb=20181107175612&path-prefix=en",
    "priestess of the wild sun": "https://static.wikia.nocookie.net/tibia/images/b/b5/Priestess_of_the_Wild_Sun.gif/revision/latest?cb=20190702012031&path-prefix=en",
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")


def parse_hunt_session(text):
    session = {}
    m = re.search(r"From (\d{4}-\d{2}-\d{2}), (\d{2}:\d{2}:\d{2}) to (\d{4}-\d{2}-\d{2}), (\d{2}:\d{2}:\d{2})", text)
    if m:
        session["date"] = m.group(1)
        session["start_time"] = f"{m.group(1)} {m.group(2)}"
        session["end_time"] = f"{m.group(3)} {m.group(4)}"

    m = re.search(r"Session:\s+(\d{2}):(\d{2})", text)
    if m:
        hours = int(m.group(1)) + int(m.group(2)) / 60
        session["hours"] = hours
        session["duration_str"] = f"{m.group(1)}:{m.group(2)}"

    fields = {
        "raw_xp_gain": r"Raw XP Gain:\s+([\d,]+)",
        "xp_gain": r"(?<!Raw )XP Gain:\s+([\d,]+)",
        "raw_xp_h": r"Raw XP/h:\s+([\d,]+)",
        "xp_h": r"(?<!Raw )XP/h:\s+([\d,]+)",
        "loot": r"Loot:\s+([\d,]+)",
        "supplies": r"Supplies:\s+([\d,]+)",
        "balance": r"Balance:\s+([\d,]+)",
        "damage": r"Damage:\s+([\d,]+)",
        "damage_h": r"Damage/h:\s+([\d,]+)",
        "healing": r"Healing:\s+([\d,]+)",
        "healing_h": r"Healing/h:\s+([\d,]+)",
    }
    for key, pattern in fields.items():
        m = re.search(pattern, text)
        if m:
            session[key] = int(m.group(1).replace(",", ""))

    def _extract_section(after, before):
        start = text.find(after)
        if start == -1:
            return ""
        start += len(after)
        end = len(text) if before is None else text.find(before, start)
        if end == -1:
            end = len(text)
        return text[start:end]

    monsters = {}
    kills_section = _extract_section("Killed Monsters:", "\nLooted Items:")
    for m in re.finditer(r"(\d+)x\s+(.+)", kills_section):
        count = int(m.group(1))
        name = m.group(2).strip()
        monsters[name] = count
    session["monsters"] = monsters
    session["total_kills"] = sum(monsters.values())

    items = {}
    loot_section = _extract_section("Looted Items:", None)
    for m in re.finditer(r"(\d+)x\s+(.+)", loot_section):
        count = int(m.group(1))
        name = m.group(2).strip()
        items[name] = count
    session["items"] = items

    if "hours" in session and session["hours"] > 0:
        session["calc_raw_xp_h"] = round(session.get("raw_xp_gain", 0) / session["hours"])
        session["calc_profit_h"] = round(session.get("balance", 0) / session["hours"])
        session["calc_xp_h"] = round(session.get("xp_gain", 0) / session["hours"])

    return session


def parse_hunt_session_json(data):
    session = {}
    session["date"] = data.get("Session start", "")[:10]
    session["start_time"] = data.get("Session start", "").replace(", ", " ")
    session["end_time"] = data.get("Session end", "").replace(", ", " ")

    dur = data.get("Session length", "")
    m = re.match(r"(\d{2}):(\d{2})", dur)
    if m:
        hours = int(m.group(1)) + int(m.group(2)) / 60
        session["hours"] = hours
        session["duration_str"] = m.group(0)

    field_map = {
        "raw_xp_gain": "Raw XP Gain",
        "xp_gain": "XP Gain",
        "raw_xp_h": "Raw XP/h",
        "xp_h": "XP/h",
        "loot": "Loot",
        "supplies": "Supplies",
        "balance": "Balance",
        "damage": "Damage",
        "damage_h": "Damage/h",
        "healing": "Healing",
        "healing_h": "Healing/h",
    }
    for key, json_key in field_map.items():
        val = data.get(json_key)
        if val is not None:
            session[key] = int(str(val).replace(",", ""))

    monsters = {}
    for entry in data.get("Killed Monsters", []):
        name = entry.get("Name", "")
        count = int(entry.get("Count", 0))
        if name:
            monsters[name] = monsters.get(name, 0) + count
    session["monsters"] = monsters
    session["total_kills"] = sum(monsters.values())

    items = {}
    for entry in data.get("Looted Items", []):
        name = entry.get("Name", "")
        count = int(entry.get("Count", 0))
        if name:
            items[name] = items.get(name, 0) + count
    session["items"] = items

    if "hours" in session and session["hours"] > 0:
        session["calc_raw_xp_h"] = round(session.get("raw_xp_gain", 0) / session["hours"])
        session["calc_profit_h"] = round(session.get("balance", 0) / session["hours"])
        session["calc_xp_h"] = round(session.get("xp_gain", 0) / session["hours"])

    return session


def _tibiawiki_name_variants(name):
    base = name.strip().title().replace(" ", "_")
    yield base
    lower_articles = re.sub(r'\b(Of|The|And|In|At|For|A|An)\b', lambda m: m.group(1).lower(), base)
    if lower_articles != base:
        yield lower_articles


def download_creature_image(name):
    url = CREATURE_IMAGES.get(name.lower())
    if not url:
        for ext in (".gif", ".png", ".jpg"):
            for variant in _tibiawiki_name_variants(name):
                safe = variant
                api_url = (
                    f"https://tibia.fandom.com/api.php?"
                    f"action=query&titles=File%3A{safe}{ext}"
                    f"&prop=imageinfo&iiprop=url&format=json"
                )
                try:
                    r = requests.get(api_url, timeout=10,
                                     headers={"User-Agent": "TibiaHuntAnalyzer/1.0"})
                    if r.status_code == 200:
                        pages = r.json().get("query", {}).get("pages", {})
                        for pid, page in pages.items():
                            if pid != "-1":
                                ii = page.get("imageinfo", [])
                                if ii:
                                    url = ii[0].get("url")
                                    break
                        if url:
                            break
                except Exception:
                    pass
            if url:
                break
    if not url:
        return None
    try:
        resp = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        if resp.status_code == 200:
            img = Image.open(io.BytesIO(resp.content))
            if img.mode in ("P", "PA"):
                img = img.convert("RGBA")
            elif img.mode != "RGBA":
                img = img.convert("RGBA")
            img = img.resize((64, 64), Image.LANCZOS)
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)
            return buf
    except Exception:
        pass
    return None


MINOR_THRESHOLD = 15


def _sessions_monsters_are_compatible(m1, m2):
    keys1 = set(m1.keys())
    keys2 = set(m2.keys())
    if keys1 == keys2:
        total1 = sum(m1.values())
        total2 = sum(m2.values())
        if total1 > 0 and total2 > 0:
            max_delta = 0
            for k in keys1:
                pct1 = m1[k] / total1 * 100
                pct2 = m2[k] / total2 * 100
                delta = abs(pct1 - pct2)
                if delta > max_delta:
                    max_delta = delta
            return max_delta < 30
        return True
    common = keys1 & keys2
    if not common:
        return False
    diff1 = [v for k, v in m1.items() if k not in common]
    diff2 = [v for k, v in m2.items() if k not in common]
    all_diffs = diff1 + diff2
    if not all_diffs:
        return True
    return max(all_diffs) < MINOR_THRESHOLD


def _clean_group_monsters(group):
    if len(group) < 2:
        for s in group:
            h = s.get("hours", 1)
            s["monsters"] = {k: v for k, v in s["monsters"].items() if v / h >= MINOR_THRESHOLD}
            s["total_kills"] = sum(s["monsters"].values())
        return
    common = None
    for s in group:
        mset = set(s["monsters"].keys())
        if common is None:
            common = mset.copy()
        else:
            common &= mset
    for s in group:
        s["monsters"] = {k: v for k, v in s["monsters"].items() if k in common}
        s["total_kills"] = sum(s["monsters"].values())


def group_sessions(sessions):
    groups = []
    used = set()
    for i, s1 in enumerate(sessions):
        if i in used:
            continue
        group = [s1]
        used.add(i)
        for j, s2 in enumerate(sessions):
            if j in used:
                continue
            if _sessions_monsters_are_compatible(
                s1["monsters"], s2["monsters"]
            ):
                group.append(s2)
                used.add(j)
        groups.append(group)
    for g in groups:
        _clean_group_monsters(g)
    return groups


def compute_group_metrics(group):
    calc_xp_h = mean(s["calc_raw_xp_h"] for s in group)
    calc_profit_h = mean(s["calc_profit_h"] for s in group)
    all_monsters = {}
    for s in group:
        for name, count in s["monsters"].items():
            all_monsters[name] = all_monsters.get(name, 0) + count
    total_hours = sum(s["hours"] for s in group)
    all_monsters = {k: v for k, v in all_monsters.items() if v / total_hours >= MINOR_THRESHOLD}
    total_kills = sum(all_monsters.values())
    all_items = {}
    for s in group:
        for name, count in s.get("items", {}).items():
            all_items[name] = all_items.get(name, 0) + count
    avg_balance = mean(s.get("balance", 0) for s in group)
    total_raw_xp = sum(s.get("raw_xp_gain", 0) for s in group)
    sorted_creatures = sorted(all_monsters.items(), key=lambda x: x[1], reverse=True)
    top_names = [c[0] for c in sorted_creatures[:1]]
    creature_name = ", ".join(top_names)
    group_name = creature_name
    return {
        "name": group_name,
        "sessions": group,
        "count": len(group),
        "avg_raw_xp_h": round(calc_xp_h),
        "avg_profit_h": round(calc_profit_h),
        "total_kills": total_kills,
        "total_hours": round(total_hours, 2),
        "avg_balance": round(avg_balance),
        "total_raw_xp": total_raw_xp,
        "monsters": sorted(all_monsters.items(), key=lambda x: x[1], reverse=True),
        "items": sorted(all_items.items(), key=lambda x: x[1], reverse=True),
    }


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(cell, alt=False):
    if alt:
        cell.fill = ALT_ROW_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")


def compute_global_kills_h(sessions):
    kills_per_creature = {}
    hours_per_creature = {}
    for s in sessions:
        h = s.get("hours", 0)
        for cname, ccount in s.get("monsters", {}).items():
            kills_per_creature[cname] = kills_per_creature.get(cname, 0) + ccount
            hours_per_creature[cname] = hours_per_creature.get(cname, 0) + h
    if not kills_per_creature:
        return []
    result = []
    for cname, ccount in kills_per_creature.items():
        total_h = hours_per_creature.get(cname, 0)
        kh = round(ccount / total_h) if total_h > 0 else 0
        result.append((cname, kh))
    result.sort(key=lambda x: x[1], reverse=True)
    return result[:5]





def apply_threshold_format(ws, group_avg, start_row, end_row, xp_col):
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=xp_col)
        val = cell.value
        if isinstance(val, (int, float)) and group_avg > 0:
            deviation = (val - group_avg) / group_avg
            if deviation < -0.30:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = RED_FILL
            elif deviation > 0.30:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = GREEN_FILL


def build_summary_sheet(wb, groups, top5=None, all_sessions=None):
    ws = wb.active
    ws.title = "Resumo"

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Tibia Hunt Analyzer - Resumo Comparativo"
    c.font = Font(bold=True, size=14, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(size=10, italic=True, color="666666")
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:G3")
    c = ws.cell(row=3, column=1)
    c.value = "TOP EXP"
    c.font = Font(bold=True, size=12, color="1F4E79")

    headers = [
        "Hunt (Criaturas)", "Sessoes", "Horas Total", "Media Raw XP/h",
        "Media Profit/h", "Total Kills", "Rank (XP/h)"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    sorted_groups = sorted(groups, key=lambda g: g["avg_raw_xp_h"], reverse=True)

    for rank, group in enumerate(sorted_groups, 1):
        row = 4 + rank
        ws.cell(row=row, column=1, value=group["name"])
        ws.cell(row=row, column=2, value=group["count"])
        h = group["total_hours"]
        ws.cell(row=row, column=3, value=f"{int(h):02d}:{round((h % 1) * 60):02d}")
        ws.cell(row=row, column=4, value=group["avg_raw_xp_h"])
        ws.cell(row=row, column=5, value=group["avg_profit_h"])
        ws.cell(row=row, column=6, value=group["total_kills"])
        ws.cell(row=row, column=7, value=f"#{rank}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            style_data_cell(cell, rank % 2 == 0)
            if col in (4, 5):
                cell.number_format = '#,##0'

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    chart_height = (len(sorted_groups) + 1) * 0.45
    chart = BarChart()
    chart.title = "XP/h por Hunt"
    chart.style = 10
    chart.width = 22
    chart.height = chart_height
    data = Reference(ws, min_col=4, min_row=4, max_row=4 + len(sorted_groups))
    cats = Reference(ws, min_col=1, min_row=5, max_row=4 + len(sorted_groups))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.x_axis.title = "Hunt"
    chart.y_axis.title = "XP/h"
    chart.y_axis.scaling.min = 0
    ws.add_chart(chart, "I4")

    section_row = 4 + len(sorted_groups) + 2

    # --- TOP 5 Profit/h (esquerda, colunas A-E) ---
    ws.merge_cells(f"A{section_row}:E{section_row}")
    c = ws.cell(row=section_row, column=1)
    c.value = "TOP 5 Profit/h"
    c.font = Font(bold=True, size=12, color="1F4E79")

    profit_headers = ["#", "Hunt", "Profit/h", "+/-% Media", "Imagem"]
    phrow = section_row + 1
    for col, h in enumerate(profit_headers, 1):
        ws.cell(row=phrow, column=col, value=h)
    style_header_row(ws, phrow, len(profit_headers))

    profit_sorted = sorted(groups, key=lambda g: g["avg_profit_h"], reverse=True)
    avg_profit = mean(g["avg_profit_h"] for g in profit_sorted[:5]) if profit_sorted else 0

    for rank, group in enumerate(profit_sorted[:5], 1):
        row = phrow + rank
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=group["name"])
        cell_p = ws.cell(row=row, column=3, value=group["avg_profit_h"])
        cell_p.number_format = '#,##0'
        if avg_profit > 0:
            pct = round((group["avg_profit_h"] - avg_profit) / avg_profit * 100, 1)
            cell_pct = ws.cell(row=row, column=4, value=f"{pct:+.1f}%")
            if pct > 0:
                cell_pct.font = Font(color="006100")
            elif pct < 0:
                cell_pct.font = Font(color="9C0006")
        else:
            ws.cell(row=row, column=4, value="-")
        for col in range(1, 6):
            style_data_cell(ws.cell(row=row, column=col), rank % 2 == 0)

        top_creature = group["monsters"][0][0]
        img_data = download_creature_image(top_creature)
        if img_data:
            xl_img = XlImage(img_data)
            xl_img.width = 64
            xl_img.height = 64
            ws.add_image(xl_img, f"E{row}")
            ws.row_dimensions[row].height = 70
        else:
            ws.cell(row=row, column=5, value="(N/D)")

    # --- Preferred List (direita, colunas G-K) ---
    if top5:
        ws.merge_cells(f"G{section_row}:K{section_row}")
        c = ws.cell(row=section_row, column=7)
        c.value = "Preferred List - Top 5 Kills/h"
        c.font = Font(bold=True, size=12, color="1F4E79")

        pl_headers = ["#", "Criatura", "Kills/h", "+/-% Media", "Imagem"]
        pl_hrow = section_row + 1
        for col, h in enumerate(pl_headers, 1):
            ws.cell(row=pl_hrow, column=col + 6, value=h)
        for col in range(7, 12):
            cell = ws.cell(row=pl_hrow, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

        if ws.column_dimensions["G"].width < 6:
            ws.column_dimensions["G"].width = 6
        if ws.column_dimensions["H"].width < 30:
            ws.column_dimensions["H"].width = 30
        if ws.column_dimensions["I"].width < 14:
            ws.column_dimensions["I"].width = 14
        if ws.column_dimensions["J"].width < 14:
            ws.column_dimensions["J"].width = 14
        if ws.column_dimensions["K"].width < 14:
            ws.column_dimensions["K"].width = 14

        avg_kh = mean(item[1] for item in top5) if top5 else 0

        for idx, (cname, kh) in enumerate(top5):
            row = pl_hrow + 1 + idx
            ws.cell(row=row, column=7, value=idx + 1)
            ws.cell(row=row, column=8, value=cname.capitalize())
            cell_kh = ws.cell(row=row, column=9, value=kh)
            cell_kh.number_format = '#,##0'
            if avg_kh > 0:
                pct = round((kh - avg_kh) / avg_kh * 100, 1)
                cell_pct = ws.cell(row=row, column=10, value=f"{pct:+.1f}%")
                if pct > 0:
                    cell_pct.font = Font(color="006100")
                elif pct < 0:
                    cell_pct.font = Font(color="9C0006")
            else:
                ws.cell(row=row, column=10, value="-")
            for col in range(7, 12):
                style_data_cell(ws.cell(row=row, column=col), idx % 2 == 0)

            img_data = download_creature_image(cname)
            if img_data:
                xl_img = XlImage(img_data)
                xl_img.width = 64
                xl_img.height = 64
                ws.add_image(xl_img, f"K{row}")
                ws.row_dimensions[row].height = 70
            else:
                ws.cell(row=row, column=11, value="(N/D)")

def build_hunt_detail_sheet(wb, group):
    top_creature = group["monsters"][0][0]
    top_name = top_creature.title()
    safe_name = re.sub(r'[\\/*?:\[\]]', '_', top_name)[:24]
    ws = wb.create_sheet(title=safe_name)

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"Hunt: {top_name}"
    c.font = Font(bold=True, size=14, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    metrics = [
        ("Sessoes agrupadas", str(group["count"])),
        ("Horas total", f"{int(group['total_hours']):02d}:{round((group['total_hours'] % 1) * 60):02d}"),
        ("Media Raw XP/h", f"{group['avg_raw_xp_h']:,}"),
        ("Media Profit/h", f"{group['avg_profit_h']:,}"),
        ("Total kills", str(group["total_kills"])),
        ("Total Raw XP", f"{group['total_raw_xp']:,}"),
    ]
    for i, (label, value) in enumerate(metrics):
        row = 3 + i
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = THIN_BORDER

    cstart = 10
    ws.merge_cells(f"A{cstart}:G{cstart}")
    ws.cell(row=cstart, column=1, value="Criaturas Mortas").font = Font(bold=True, size=12, color="1F4E79")

    creature_headers = ["", "Criatura", "Kills", "Kills/h", "% Total", "Imagem"]
    chrow = cstart + 1
    for col, h in enumerate(creature_headers, 1):
        ws.cell(row=chrow, column=col, value=h)
    style_header_row(ws, chrow, len(creature_headers))

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 18

    for idx, (cname, ccount) in enumerate(group["monsters"]):
        row = chrow + 1 + idx
        ws.cell(row=row, column=1, value=idx + 1)
        ws.cell(row=row, column=2, value=cname.capitalize())
        ws.cell(row=row, column=3, value=ccount)
        kills_h = round(ccount / group["total_hours"]) if group["total_hours"] > 0 else 0
        ws.cell(row=row, column=4, value=kills_h)
        ws.cell(row=row, column=4).number_format = '#,##0'
        pct = round(ccount / group["total_kills"] * 100, 1) if group["total_kills"] > 0 else 0
        ws.cell(row=row, column=5, value=f"{pct}%")
        for col in range(1, 7):
            style_data_cell(ws.cell(row=row, column=col), idx % 2 == 0)

        img_data = download_creature_image(cname)
        if img_data:
            xl_img = XlImage(img_data)
            xl_img.width = 64
            xl_img.height = 64
            ws.add_image(xl_img, f"F{row}")
            ws.row_dimensions[row].height = 70
            ws.column_dimensions["F"].width = 12
        else:
            ws.cell(row=row, column=6, value="(imagem N/D)")

    sstart = cstart + len(group["monsters"]) + 3

    sessions_with_date = [(s.get("date", ""), s.get("calc_raw_xp_h", 0)) for s in group["sessions"] if s.get("date")]
    sessions_with_date.sort(key=lambda x: x[0])

    if len(sessions_with_date) >= 2:
        mid = len(sessions_with_date) // 2
        first_half = [x[1] for x in sessions_with_date[:mid]]
        second_half = [x[1] for x in sessions_with_date[mid:]]
        avg_first = mean(first_half) if first_half else 0
        avg_second = mean(second_half) if second_half else 0

        if avg_first > 0:
            ratio = avg_second / avg_first
            if ratio >= 1.05:
                trend = "Melhorando"
                trend_color = "006100"
            elif ratio <= 0.95:
                trend = "Piorando"
                trend_color = "9C0006"
            else:
                trend = "Estavel"
                trend_color = "666666"
        else:
            trend = "-"
            trend_color = "666666"

        ts = sstart
        ws.merge_cells(f"A{ts}:F{ts}")
        c = ws.cell(row=ts, column=1, value="Tendencia Historica")
        c.font = Font(bold=True, size=12, color="1F4E79")

        trend_labels = [
            ("Tendencia", trend),
            ("Media 1a metade", f"{avg_first:,.0f} XP/h"),
            ("Media 2a metade", f"{avg_second:,.0f} XP/h"),
            ("Melhor sessao", f"{max(x[1] for x in sessions_with_date):,} XP/h ({max(sessions_with_date, key=lambda x: x[1])[0]})"),
            ("Pior sessao", f"{min(x[1] for x in sessions_with_date):,} XP/h ({min(sessions_with_date, key=lambda x: x[1])[0]})"),
        ]
        for i, (label, value) in enumerate(trend_labels):
            row = ts + 1 + i
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=1).border = THIN_BORDER
            cv = ws.cell(row=row, column=2, value=value)
            cv.border = THIN_BORDER
            if label == "Tendencia":
                cv.font = Font(color=trend_color, bold=True)

        sstart = ts + len(trend_labels) + 2

    ws.merge_cells(f"A{sstart}:I{sstart}")
    ws.cell(row=sstart, column=1, value="Detalhes das Sessoes neste Grupo").font = Font(bold=True, size=12, color="1F4E79")

    sess_headers = ["Sessao", "Duracao", "Raw XP Gain", "Raw XP/h", "Profit/h", "Balance", "Kills", "Damage", "Damage/h", "% Media"]
    shrow = sstart + 1
    for col, h in enumerate(sess_headers, 1):
        ws.cell(row=shrow, column=col, value=h)
    style_header_row(ws, shrow, len(sess_headers))

    group_avg = group["avg_raw_xp_h"]
    for idx, s in enumerate(group["sessions"]):
        row = shrow + 1 + idx
        ws.cell(row=row, column=1, value=s.get("source", f"Sessao {idx+1}"))
        ws.cell(row=row, column=2, value=s.get("duration_str", ""))
        ws.cell(row=row, column=3, value=s.get("raw_xp_gain", 0))
        ws.cell(row=row, column=4, value=s.get("calc_raw_xp_h", 0))
        ws.cell(row=row, column=5, value=s.get("calc_profit_h", 0))
        ws.cell(row=row, column=6, value=s.get("balance", 0))
        ws.cell(row=row, column=7, value=s.get("total_kills", 0))
        ws.cell(row=row, column=8, value=s.get("damage", 0))
        dmg_h = round(s.get("damage", 0) / s.get("hours", 1)) if s.get("hours", 0) > 0 else 0
        ws.cell(row=row, column=9, value=dmg_h)
        session_xp = s.get("calc_raw_xp_h", 0)
        if group_avg > 0:
            dev = round((session_xp - group_avg) / group_avg * 100, 1)
            ws.cell(row=row, column=10, value=f"{dev:+.1f}%")
        else:
            ws.cell(row=row, column=10, value="-")
        for col in range(1, len(sess_headers) + 1):
            cell = ws.cell(row=row, column=col)
            style_data_cell(cell, idx % 2 == 0)
            if col >= 3:
                cell.number_format = '#,##0'

    n_sessions = len(group["sessions"])
    if n_sessions > 0:
        apply_threshold_format(ws, group_avg, shrow + 1, shrow + n_sessions, 4)

        label_col = 13
        ws.cell(row=shrow, column=label_col, value="Data")
        for idx, s in enumerate(group["sessions"]):
            raw_date = s.get("date", "")
            label = raw_date[-5:] if len(raw_date) >= 10 else s.get("source", f"S{idx+1}")[:10]
            ws.cell(row=shrow + 1 + idx, column=label_col, value=label)
        ws.column_dimensions[get_column_letter(label_col)].hidden = True

        chart = LineChart()
        chart.title = "XP/h | Profit/h | Dano/h por Sessao"
        chart.style = 10
        chart.width = 22
        chart.height = 12
        xp_data = Reference(ws, min_col=4, min_row=shrow, max_row=shrow + n_sessions)
        profit_data = Reference(ws, min_col=5, min_row=shrow, max_row=shrow + n_sessions)
        dmg_data = Reference(ws, min_col=9, min_row=shrow, max_row=shrow + n_sessions)
        scats = Reference(ws, min_col=label_col, min_row=shrow + 1, max_row=shrow + n_sessions)
        chart.add_data(xp_data, titles_from_data=True)
        chart.add_data(profit_data, titles_from_data=True)
        chart.add_data(dmg_data, titles_from_data=True)
        chart.set_categories(scats)
        chart.x_axis.title = "Sessao"
        chart.y_axis.title = "XP/h | Profit/h | Dano/h"
        chart.y_axis.numFmt = '#,##0'
        chart.y_axis.scaling.min = 0
        ws.add_chart(chart, f"K1")

    all_cnames = [cname for cname, _ in group["monsters"]]
    km_start = shrow + len(group["sessions"]) + 3
    ws.merge_cells(f"A{km_start}:{get_column_letter(4 + len(all_cnames))}{km_start}")
    ws.cell(row=km_start, column=1, value="Kills por Sessao / Criatura").font = Font(bold=True, size=12, color="1F4E79")

    km_headers = ["Sessao"] + [c.capitalize() for c in all_cnames] + ["Total"]
    km_hrow = km_start + 1
    for col, h in enumerate(km_headers, 1):
        ws.cell(row=km_hrow, column=col, value=h)
    style_header_row(ws, km_hrow, len(km_headers))

    for idx, s in enumerate(group["sessions"]):
        row = km_hrow + 1 + idx
        ws.cell(row=row, column=1, value=s.get("source", f"Sessao {idx+1}"))
        row_total = 0
        for ci, cname in enumerate(all_cnames):
            k = s.get("monsters", {}).get(cname, 0)
            ws.cell(row=row, column=2 + ci, value=k)
            row_total += k
        ws.cell(row=row, column=2 + len(all_cnames), value=row_total)
        for col in range(1, len(km_headers) + 1):
            cell = ws.cell(row=row, column=col)
            style_data_cell(cell, idx % 2 == 0)

    for ci in range(len(all_cnames)):
        col_letter = get_column_letter(2 + ci)
        if ws.column_dimensions[col_letter].width < 18:
            ws.column_dimensions[col_letter].width = 18

    all_icnames = [iname for iname, _ in group.get("items", [])]
    if all_icnames:
        istart = km_hrow + len(group["sessions"]) + 3
        ws.merge_cells(f"A{istart}:D{istart}")
        ws.cell(row=istart, column=1, value="Itens Looteados").font = Font(bold=True, size=12, color="1F4E79")

        iheaders = ["", "Item", "Qty", "Qty/h"]
        ihrow = istart + 1
        for col, h in enumerate(iheaders, 1):
            ws.cell(row=ihrow, column=col, value=h)
        style_header_row(ws, ihrow, len(iheaders))

        if ws.column_dimensions["A"].width < 6:
            ws.column_dimensions["A"].width = 6
        if ws.column_dimensions["B"].width < 40:
            ws.column_dimensions["B"].width = 40

        for idx, (iname, icount) in enumerate(group["items"]):
            row = ihrow + 1 + idx
            ws.cell(row=row, column=1, value=idx + 1)
            ws.cell(row=row, column=2, value=iname)
            ws.cell(row=row, column=3, value=icount)
            ih = round(icount / group["total_hours"]) if group["total_hours"] > 0 else 0
            ws.cell(row=row, column=4, value=ih)
            ws.cell(row=row, column=4).number_format = '#,##0'
            for col in range(1, len(iheaders) + 1):
                style_data_cell(ws.cell(row=row, column=col), idx % 2 == 0)


def split_multi_session(text, base_name):
    count = text.count("Session data:")
    if count <= 1:
        return [(text, base_name)]
    chunks = [c.strip() for c in text.split("\n\n") if c.strip() and "Session data:" in c]
    result = []
    for i, chunk in enumerate(chunks, 1):
        suffix = f" (#{i})" if count > 1 else ""
        result.append((chunk, f"{base_name}{suffix}"))
    return result


def _script_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _tibia_log_dir():
    local_appdata = os.environ.get("LOCALAPPDATA", "")
    if not local_appdata:
        return None
    p = Path(local_appdata) / "Tibia" / "packages" / "Tibia" / "log"
    return p if p.is_dir() else None


def load_json_sessions(tibia_dir):
    entries = []
    files = sorted(tibia_dir.glob("Hunting_Session_*.json"))
    if not files:
        files = sorted(tibia_dir.glob("*.json"))
    if not files:
        print(f"[ERRO] Nenhum arquivo .json encontrado em: {tibia_dir}")
        return None
    for f in files:
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            s = parse_hunt_session_json(data)
            if not s or "hours" not in s:
                print(f"[AVISO] '{f.stem}': dados de sessao invalidos no JSON")
                continue
            s["source"] = f.stem
            entries.append(s)
        except (json.JSONDecodeError, KeyError, ValueError) as e:
            print(f"[AVISO] '{f.name}': erro ao ler JSON ({e})")
            continue
    return entries


def _parse_txt_sessions(raw_entries):
    sessions = []
    for text, source in raw_entries:
        s = parse_hunt_session(text)
        if not s or "hours" not in s:
            print(f"[AVISO] '{source}': nao foi possivel extrair dados validos")
            return None
        s["source"] = source
        sessions.append(s)
    return sessions


def load_sessions(args):
    if args.dir:
        p = Path(args.dir)
        if not p.is_dir():
            print(f"[ERRO] Diretorio nao encontrado: {args.dir}")
            return None
        files = sorted(p.glob("*.txt")) + sorted(p.glob("*.log"))
        if not files:
            print(f"[ERRO] Nenhum arquivo .txt ou .log em: {args.dir}")
            return None
        raw = []
        for f in files:
            raw.extend(split_multi_session(f.read_text(encoding="utf-8"), f.stem))
        print(f"  Carregados {len(raw)} sessao(oes) de {len(files)} arquivo(s) em: {args.dir}")
        return _parse_txt_sessions(raw)

    if args.file:
        raw = []
        for fpath in args.file:
            if not os.path.exists(fpath):
                print(f"[ERRO] Arquivo nao encontrado: {fpath}")
                return None
            fname = Path(fpath).stem
            with open(fpath, "r", encoding="utf-8") as fh:
                raw.extend(split_multi_session(fh.read(), fname))
        print(f"  Carregados {len(raw)} sessao(oes) via --file")
        return _parse_txt_sessions(raw)

    if args.paste:
        print("Cole todas as sessoes (separadas por linhas em branco). Pressione Ctrl+Z + Enter (Windows) ou Ctrl+D (Linux/Mac) para finalizar:")
        stdin_data = sys.stdin.read()
        chunks = [c.strip() for c in stdin_data.split("\n\n") if c.strip()]
        if len(chunks) < 1:
            print("[ERRO] Nenhum dado encontrado.")
            return None
        raw = [(chunk, f"Sessao {i}") for i, chunk in enumerate(chunks, 1)]
        print(f"  Lidas {len(raw)} sessoes do stdin")
        return _parse_txt_sessions(raw)

    tibia_dir = _tibia_log_dir()
    if tibia_dir is not None:
        json_files = sorted(tibia_dir.glob("Hunting_Session_*.json"))
        if json_files:
            return load_json_sessions(tibia_dir)
    hunts_dir = _script_dir() / "hunts"
    if hunts_dir.is_dir() and list(hunts_dir.glob("*.txt")):
        files = sorted(hunts_dir.glob("*.txt"))
        raw = []
        for f in files:
            raw.extend(split_multi_session(f.read_text(encoding="utf-8"), f.stem))
        print(f"  Carregados {len(raw)} sessao(oes) de {len(files)} arquivo(s) em: {hunts_dir}")
        return _parse_txt_sessions(raw)
    if tibia_dir is None:
        print(f"[ERRO] Pasta de log do Tibia nao encontrada.")
    else:
        print(f"[ERRO] Nenhum arquivo JSON na pasta de log do Tibia nem .txt em: {hunts_dir}")
    return None


def main():
    parser = argparse.ArgumentParser(description="Tibia Hunt Analyzer Excel Generator")
    parser.add_argument("--file", "-f", action="append", dest="file",
                        help="Arquivo de sessao (pode repetir: -f s1.txt -f s2.txt ...)")
    parser.add_argument("--dir", "-d",
                        help="Diretorio com arquivos .txt/.log de sessoes")
    parser.add_argument("-o", "--output", help="Arquivo Excel de saida")
    parser.add_argument("--paste", action="store_true",
                        help="Ler sessoes do teclado (cole os dados)")
    args = parser.parse_args()

    sessions = load_sessions(args)
    if sessions is None:
        return

    print("=" * 60)
    print(f"Tibia Hunt Analyzer - Excel Generator ({len(sessions)} sessoes)")
    print("=" * 60)

    for s in sessions:
        print(f"\n--- {s['source']} ---")
        print(f"  Duracao: {s.get('duration_str', 'N/A')} ({s.get('hours', 0):.2f}h)")
        print(f"  Raw XP Gain: {s.get('raw_xp_gain', 0):,}")
        print(f"  Raw XP/h (calc): {s.get('calc_raw_xp_h', 0):,}")
        print(f"  Balance: {s.get('balance', 0):,}")
        print(f"  Profit/h (calc): {s.get('calc_profit_h', 0):,}")
        print(f"  Total kills: {s.get('total_kills', 0)}")
        for cname, ccount in s.get("monsters", {}).items():
            print(f"    {ccount}x {cname}")

    groups = group_sessions(sessions)
    print(f"\n--- Agrupamento: {len(groups)} grupo(s) ---")
    group_metrics = []
    for g in groups:
        gm = compute_group_metrics(g)
        group_metrics.append(gm)
        print(f"\n  Grupo: {gm['name']} ({gm['count']} sessoes)")
        print(f"  Media Raw XP/h: {gm['avg_raw_xp_h']:,}  |  Profit/h: {gm['avg_profit_h']:,}")
        for cname, ccount in gm["monsters"]:
            print(f"    {ccount}x {cname}")

    wb = Workbook()
    top5 = compute_global_kills_h(sessions)
    build_summary_sheet(wb, group_metrics, top5, sessions)

    for gm in sorted(group_metrics, key=lambda g: g["avg_raw_xp_h"], reverse=True):
        build_hunt_detail_sheet(wb, gm)

    output_path = args.output or "hunt_analysis.xlsx"
    wb.save(output_path)
    print(f"\n[OK] Planilha salva: {output_path}")
    print(f"Arquivo: {Path(output_path).absolute()}")

    if sys.stdout.isatty() and sys.platform == "win32":
        input("\nPressione ENTER para sair...")


if __name__ == "__main__":
    main()
